from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
ERROR_FILL = PatternFill("solid", fgColor="FDE9E7")
THIN_BORDER = Border(bottom=Side(style="thin", color="D0D0D0"))



def autosize_columns(ws) -> None:
    for column in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[col_letter].width = min(max(max_length + 2, 12), 60)



def style_header(ws, row: int) -> None:
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER



def write_workbook(
    output_path: Path,
    summary_rows: list[dict],
    evidence_rows: list[dict],
    clean_rows: list[dict],
    metadata: dict,
    stage_logs: list[dict],
) -> None:
    wb = Workbook()
    ws_dq = wb.active
    ws_dq.title = "Data_Quality"

    ws_dq["A1"] = "Data Quality Report"
    ws_dq["A1"].font = Font(size=16, bold=True)

    metrics_rows = [
        ("Source file(s)", metadata["source_file"]),
        ("Source count", metadata["source_count"]),
        ("Generated at", metadata["generated_at"]),
        ("Total rows", metadata["total_rows"]),
        ("Rows with issues", metadata["rows_with_issues"]),
        ("Total evidences", metadata["total_evidences"]),
        ("DQ score", metadata["dq_score"]),
        ("Failed rules", metadata["failed_rules"]),
    ]

    for row_idx, (label, value) in enumerate(metrics_rows, start=3):
        ws_dq.cell(row=row_idx, column=1, value=label)
        ws_dq.cell(row=row_idx, column=2, value=value)
        ws_dq.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_dq.cell(row=row_idx, column=1).fill = SECTION_FILL

    ws_dq["B9"].number_format = "0.0%"

    ws_dq["A12"] = "Summary by rule"
    ws_dq["A12"].font = Font(size=12, bold=True)
    start_row = 13
    headers = ["rule_name", "severity", "description", "failed_rows", "pass_rate"]
    for index, header in enumerate(headers, start=1):
        ws_dq.cell(row=start_row, column=index, value=header)
    style_header(ws_dq, start_row)

    for row_index, item in enumerate(summary_rows, start=start_row + 1):
        for col_index, header in enumerate(headers, start=1):
            ws_dq.cell(row=row_index, column=col_index, value=item[header])
        ws_dq.cell(row=row_index, column=5).number_format = "0.0%"

    for row in range(start_row + 1, start_row + 1 + len(summary_rows)):
        if ws_dq[f"B{row}"].value == "high":
            for col in range(1, 6):
                ws_dq.cell(row=row, column=col).fill = ERROR_FILL

    logs_title_row = start_row + len(summary_rows) + 3
    ws_dq.cell(row=logs_title_row, column=1, value="Pipeline checkpoints")
    ws_dq.cell(row=logs_title_row, column=1).font = Font(size=12, bold=True)

    log_header_row = logs_title_row + 1
    log_headers = ["stage", "checkpoint", "details"]
    for index, header in enumerate(log_headers, start=1):
        ws_dq.cell(row=log_header_row, column=index, value=header)
    style_header(ws_dq, log_header_row)

    for row_index, item in enumerate(stage_logs, start=log_header_row + 1):
        for col_index, header in enumerate(log_headers, start=1):
            ws_dq.cell(row=row_index, column=col_index, value=item[header])

    ws_ev = wb.create_sheet("Evidences")
    ev_headers = ["row_number", "rule_name", "severity", "column_name", "message", "invalid_value", "source_file"]
    for index, header in enumerate(ev_headers, start=1):
        ws_ev.cell(row=1, column=index, value=header)
    style_header(ws_ev, 1)

    for row_index, item in enumerate(evidence_rows, start=2):
        for col_index, header in enumerate(ev_headers, start=1):
            ws_ev.cell(row=row_index, column=col_index, value=item.get(header))

    ws_clean = wb.create_sheet("Clean_Data")
    if clean_rows:
        clean_headers = list(clean_rows[0].keys())
        for index, header in enumerate(clean_headers, start=1):
            ws_clean.cell(row=1, column=index, value=header)
        style_header(ws_clean, 1)
        for row_index, item in enumerate(clean_rows, start=2):
            for col_index, header in enumerate(clean_headers, start=1):
                ws_clean.cell(row=row_index, column=col_index, value=item[header])

    for ws in wb.worksheets:
        autosize_columns(ws)
        ws.freeze_panes = "A2" if ws.title != "Data_Quality" else "A13"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
