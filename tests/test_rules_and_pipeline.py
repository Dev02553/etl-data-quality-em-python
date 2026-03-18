from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from etl_dq.pipeline import run_pipeline
from etl_dq.rules import build_rules


# ─── Helpers ────────────────────────────────────────────────────────────────

def _base_df(**overrides) -> pd.DataFrame:
    """DataFrame válido base — sobrescreva colunas individualmente nos testes."""
    data = {
        "customer_id": ["1", "2"],
        "email": ["a@example.com", "b@example.com"],
        "country": ["BR", "CL"],
        "signup_date": ["2024-01-01", "2024-01-01"],
        "credit_limit": [100, 200],
    }
    data.update(overrides)
    return pd.DataFrame(data)


def _get_rule(name: str):
    return next(r for r in build_rules() if r.name == name)


# ─── Testes de regras existentes ────────────────────────────────────────────

def test_email_format_rule_flags_invalid_values():
    df = _base_df(email=["ok@example.com", "invalid-email"])
    mask = _get_rule("email_format").evaluator(df)
    assert mask.tolist() == [False, True]


def test_unique_customer_id_rule_flags_duplicates():
    df = _base_df(customer_id=["A100", "A100"])
    mask = _get_rule("unique_customer_id").evaluator(df)
    assert mask.tolist() == [True, True]


def test_type_rules_flag_invalid_date_and_non_numeric_values():
    df = _base_df(signup_date=["2024-01-01", "not-a-date"], credit_limit=[100, "abc"])
    assert _get_rule("signup_date_valid_type").evaluator(df).tolist() == [False, True]
    assert _get_rule("credit_limit_numeric_type").evaluator(df).tolist() == [False, True]


# ─── Novos testes de regras ──────────────────────────────────────────────────

def test_required_customer_id_flags_blank():
    df = _base_df(customer_id=["", "2"])
    mask = _get_rule("required_customer_id").evaluator(df)
    assert mask.tolist() == [True, False]


def test_required_email_flags_blank():
    df = _base_df(email=["", "b@example.com"])
    mask = _get_rule("required_email").evaluator(df)
    assert mask.tolist() == [True, False]


def test_allowed_country_flags_invalid_country():
    df = _base_df(country=["BR", "XX"])
    mask = _get_rule("allowed_country").evaluator(df)
    assert mask.tolist() == [False, True]


def test_allowed_country_does_not_flag_blank():
    """Blank deve ser ignorado por allowed_country — já coberto por required."""
    df = _base_df(country=["BR", ""])
    mask = _get_rule("allowed_country").evaluator(df)
    assert mask.tolist() == [False, False]


def test_credit_limit_non_negative_flags_negative():
    df = _base_df(credit_limit=[100, -1])
    mask = _get_rule("credit_limit_non_negative").evaluator(df)
    assert mask.tolist() == [False, True]


def test_signup_date_not_future_flags_future_date():
    df = _base_df(signup_date=["2024-01-01", "2099-01-01"])
    mask = _get_rule("signup_date_not_future").evaluator(df)
    assert mask.tolist() == [False, True]


# ─── Testes de pipeline ──────────────────────────────────────────────────────

def test_pipeline_generates_workbook_from_directory(tmp_path: Path):
    input_dir = tmp_path / "data"
    input_dir.mkdir()
    output_path = tmp_path / "report.xlsx"
    (input_dir / "input_a.csv").write_text(
        "\n".join([
            "customer_id,email,country,signup_date,credit_limit",
            "1,valid@example.com,BR,2024-01-01,100",
            "1,invalid-email,XX,2099-01-01,-10",
        ]),
        encoding="utf-8",
    )

    result = run_pipeline(input_dir, output_path)
    assert output_path.exists()
    assert result["metadata"]["source_count"] == 1
    assert result["metadata"]["total_rows"] == 2
    assert result["metadata"]["rows_with_issues"] == 2
    assert result["metadata"]["total_evidences"] >= 4

    wb = load_workbook(output_path)
    assert wb.sheetnames == ["Data_Quality", "Evidences", "Clean_Data"]
    assert wb["Data_Quality"]["A1"].value == "Data Quality Report"


def test_pipeline_raises_on_missing_columns(tmp_path: Path):
    """Pipeline deve lançar ValueError com mensagem clara se coluna obrigatória faltar."""
    input_dir = tmp_path / "data"
    input_dir.mkdir()
    output_path = tmp_path / "report.xlsx"
    (input_dir / "input_bad.csv").write_text(
        "\n".join([
            "customer_id,email",
            "1,a@example.com",
        ]),
        encoding="utf-8",
    )

    with pytest.raises(ValueError, match="Colunas ausentes no CSV"):
        run_pipeline(input_dir, output_path)


def test_pipeline_raises_on_empty_directory(tmp_path: Path):
    """Pipeline deve lançar FileNotFoundError se a pasta não tiver CSVs."""
    empty_dir = tmp_path / "empty"
    empty_dir.mkdir()
    output_path = tmp_path / "report.xlsx"

    with pytest.raises(FileNotFoundError):
        run_pipeline(empty_dir, output_path)